"""
One-time script: Parse SMA Excel files → questions/warehouse.json and questions/retail.json
Extracts: question text, level, audit_methods, answered_by, standard (system pillar)
Run: python3 extract_questions.py
"""
import json
import re
import warnings
import openpyxl
from pathlib import Path

warnings.filterwarnings("ignore")

SOURCE_DIR = Path("/Users/ktp/Desktop/Safety_Career_Skills/_Source_References")
WH_EXCEL = SOURCE_DIR / "[ENG] Safety maturity assessment checklist.xlsx"
RETAIL_EXCEL = SOURCE_DIR / "[ENG] Safety Maturity Assessment Sheet for Retail.xlsx"
OUT_DIR = Path("questions")
OUT_DIR.mkdir(exist_ok=True)

LEVEL_MAP = {
    "ad-hoc": 1, "ad hoc": 1, "adhoc": 1,
    "reactive": 2,
    "standardized": 3, "standard": 3,
    "proactive": 4,
    "excellent": 5, "excellence": 5,
}
CHECK_MARK = "レ"


def clean(text):
    if not text:
        return ""
    text = str(text).strip().split("\n")[0].strip()
    return text.strip()


def clean_answered_by(text):
    if not text:
        return ""
    text = str(text).strip()
    # Remove Japanese lines, keep first English line
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    # Filter out Japanese-only lines (lines with mostly CJK characters)
    english_lines = [l for l in lines if not all(ord(c) > 127 or c in " .,/-&()" for c in l)]
    if english_lines:
        return english_lines[0].strip().rstrip("\\").strip()
    return lines[0] if lines else ""


def parse_level(cell_val):
    if not cell_val:
        return None
    v = str(cell_val).strip().lower()
    for k, lv in LEVEL_MAP.items():
        if k in v:
            return lv
    return None


def slug(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def get_methods(interview_val, onsite_val, doc_val):
    methods = []
    if interview_val and str(interview_val).strip() == CHECK_MARK:
        methods.append("interview")
    if onsite_val and str(onsite_val).strip() == CHECK_MARK:
        methods.append("onsite")
    if doc_val and str(doc_val).strip() == CHECK_MARK:
        methods.append("document")
    return methods


def extract_warehouse():
    wb = openpyxl.load_workbook(WH_EXCEL, data_only=True)
    pillars = []

    # Leadership/TM Engagement/Organization
    # B=Level(1), C=Element(2), D=No(3), E=Question(4), F=AnsweredBy(5),
    # G=Interview(6), H=Onsite(7), I=Document(8)
    for pillar_name, sheet_name, pillar_id, q_prefix in [
        ("Leadership", "Leadership", "leadership", "LD"),
        ("Teammate Engagement", "TM Engagement", "tm_engagement", "TM"),
        ("Organization", "Organization", "organization", "OR"),
    ]:
        ws = wb[sheet_name]
        elements_dict = {}
        element_order = []
        current_level = None
        last_element = None

        for row in ws.iter_rows(min_row=3, values_only=True):
            level_val = row[1]
            element_val = row[2]
            no_val = row[3]
            question_val = row[4]
            answered_by_val = row[5]
            interview_val = row[6]
            onsite_val = row[7]
            doc_val = row[8]

            lv = parse_level(level_val)
            if lv:
                current_level = lv

            if no_val is not None and str(no_val).strip().isdigit():
                element_name = clean(element_val) or last_element or "General"
                last_element = element_name

                if element_name not in elements_dict:
                    elements_dict[element_name] = []
                    element_order.append(element_name)

                elements_dict[element_name].append({
                    "id": f"WH-{q_prefix}-{int(no_val):03d}",
                    "no": int(no_val),
                    "text": clean(question_val),
                    "level": current_level or 1,
                    "answered_by": clean_answered_by(answered_by_val),
                    "audit_methods": get_methods(interview_val, onsite_val, doc_val),
                    "standard": None,
                })

        elements_list = [
            {"id": slug(n), "name": n, "questions": elements_dict[n]}
            for n in element_order
        ]
        pillars.append({"id": pillar_id, "name": pillar_name, "elements": elements_list})

    # System
    # B=Level(1), C=Standard(2), D=Element(3), E=No(4), F=Question(5),
    # G=AnsweredBy(6), H=Interview(7), I=Onsite(8), J=Document(9)
    ws = wb["System"]
    standards_dict = {}
    standard_order = []
    current_level = None
    last_standard = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        level_val = row[1]
        standard_val = row[2]
        no_val = row[4]
        question_val = row[5]
        answered_by_val = row[6]
        interview_val = row[7]
        onsite_val = row[8]
        doc_val = row[9]

        lv = parse_level(level_val)
        if lv:
            current_level = lv

        if no_val is not None and str(no_val).strip().isdigit():
            std_name = clean(standard_val) or last_standard or "General"
            last_standard = std_name

            if std_name not in standards_dict:
                standards_dict[std_name] = []
                standard_order.append(std_name)

            standards_dict[std_name].append({
                "id": f"WH-SY-{int(no_val):03d}",
                "no": int(no_val),
                "text": clean(question_val),
                "level": current_level or 1,
                "answered_by": clean_answered_by(answered_by_val),
                "audit_methods": get_methods(interview_val, onsite_val, doc_val),
                "standard": std_name,
            })

    elements_list = [
        {"id": slug(n), "name": n, "questions": standards_dict[n]}
        for n in standard_order
    ]
    pillars.append({"id": "system", "name": "System", "elements": elements_list})

    return {"type": "warehouse", "pillars": pillars}


def extract_retail():
    wb = openpyxl.load_workbook(RETAIL_EXCEL, data_only=True)
    result = {}

    for scope, scope_id in [("Store", "store"), ("Local company", "local_company")]:
        pillars = []

        # Non-system: A=Level(0), D=Element(3), E=No(4), F=Question(5), H=AnsweredBy(7)
        # Retail doesn't have レ audit method markers — infer from answered_by
        for pillar_name, sheet_name, pillar_id, q_prefix in [
            ("Leadership", f"Leadership - {scope}", "leadership", "LD"),
            ("Teammate Engagement", f"TM Engagement - {scope}", "tm_engagement", "TM"),
            ("Organization", f"Organization - {scope}", "organization", "OR"),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            elements_dict = {}
            element_order = []
            current_level = None
            last_element = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                level_val = row[0]
                element_val = row[3]
                no_val = row[4]
                question_val = row[5]
                answered_by_val = row[7]  # column H

                lv = parse_level(level_val)
                if lv:
                    current_level = lv

                if no_val is not None and str(no_val).strip().isdigit():
                    element_name = clean(element_val) or last_element or "General"
                    last_element = element_name

                    if element_name not in elements_dict:
                        elements_dict[element_name] = []
                        element_order.append(element_name)

                    sc_prefix = "ST" if scope_id == "store" else "LC"
                    ab = clean_answered_by(answered_by_val)
                    # Infer method: if "Tour" in answered_by → onsite; else interview
                    methods = ["onsite"] if "tour" in ab.lower() else ["interview"]

                    elements_dict[element_name].append({
                        "id": f"RT-{sc_prefix}-{q_prefix}-{int(no_val):03d}",
                        "no": int(no_val),
                        "text": clean(question_val),
                        "level": current_level or 1,
                        "answered_by": ab,
                        "audit_methods": methods,
                        "standard": None,
                    })

            elements_list = [
                {"id": slug(n), "name": n, "questions": elements_dict[n]}
                for n in element_order
            ]
            pillars.append({"id": pillar_id, "name": pillar_name, "elements": elements_list})

        # System: A=Level(0), D=Standard(3), E=Element(4), F=No(5), G=Question(6), I=AnsweredBy(8)
        sheet_name = f"System - {scope}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            standards_dict = {}
            standard_order = []
            current_level = None
            last_standard = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                level_val = row[0]
                standard_val = row[3]
                no_val = row[5]
                question_val = row[6]
                answered_by_val = row[8]  # column I

                lv = parse_level(level_val)
                if lv:
                    current_level = lv

                if no_val is not None and str(no_val).strip().isdigit():
                    std_name = clean(standard_val) or last_standard or "General"
                    last_standard = std_name

                    if std_name not in standards_dict:
                        standards_dict[std_name] = []
                        standard_order.append(std_name)

                    sc_prefix = "ST" if scope_id == "store" else "LC"
                    ab = clean_answered_by(answered_by_val)
                    methods = ["onsite"] if "tour" in ab.lower() else ["interview", "document"]

                    standards_dict[std_name].append({
                        "id": f"RT-{sc_prefix}-SY-{int(no_val):03d}",
                        "no": int(no_val),
                        "text": clean(question_val),
                        "level": current_level or 1,
                        "answered_by": ab,
                        "audit_methods": methods,
                        "standard": std_name,
                    })

            elements_list = [
                {"id": slug(n), "name": n, "questions": standards_dict[n]}
                for n in standard_order
            ]
            pillars.append({"id": "system", "name": "System", "elements": elements_list})

        result[scope_id] = {"scope": scope, "pillars": pillars}

    return {"type": "retail", "scopes": result}


if __name__ == "__main__":
    print("Extracting Warehouse SMA questions...")
    wh_data = extract_warehouse()
    total_q = sum(len(q["questions"]) for p in wh_data["pillars"] for q in p["elements"])
    for p in wh_data["pillars"]:
        n = sum(len(e["questions"]) for e in p["elements"])
        print(f"  {p['name']}: {n} questions")
        # Sample audit methods
        for e in p["elements"][:1]:
            for q in e["questions"][:2]:
                print(f"    Q{q['no']}: methods={q['audit_methods']} by={q['answered_by'][:30]}")
    print(f"  Total: {total_q}")
    with open(OUT_DIR / "warehouse.json", "w", encoding="utf-8") as f:
        json.dump(wh_data, f, indent=2, ensure_ascii=False)
    print(f"  Saved → questions/warehouse.json")

    print("\nExtracting Retail SMA questions...")
    rt_data = extract_retail()
    for scope_id, scope_data in rt_data["scopes"].items():
        total = sum(len(q["questions"]) for p in scope_data["pillars"] for q in p["elements"])
        print(f"  {scope_data['scope']}: {total} questions")
    with open(OUT_DIR / "retail.json", "w", encoding="utf-8") as f:
        json.dump(rt_data, f, indent=2, ensure_ascii=False)
    print(f"  Saved → questions/retail.json\nDone.")
