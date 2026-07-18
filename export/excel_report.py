"""Generate Excel assessment report using openpyxl."""
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

LEVEL_NAMES = {1: "Ad-hoc", 2: "Reactive", 3: "Standardized", 4: "Proactive", 5: "Excellence"}
LEVEL_HEX = {1: "DC3545", 2: "FD7E14", 3: "0DCAF0", 4: "0D6EFD", 5: "198754"}
ANSWER_HEX = {
    "yes": "D1E7DD", "no": "F8D7DA",
    "na": "E9ECEF", "not_rolled_out": "FFF3CD",
}
ANSWER_LABELS = {"yes": "Yes", "no": "No", "na": "N/A", "not_rolled_out": "Not rolled out", "": "—"}

THIN = Side(style="thin", color="DEE2E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DARK_FILL = PatternFill("solid", fgColor="212529")
LIGHT_FILL = PatternFill("solid", fgColor="F8F9FA")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

def level_fill(score):
    hex_c = LEVEL_HEX.get(round(score) if score else 0, "6C757D")
    return PatternFill("solid", fgColor=hex_c)

def set_cell(ws, row, col, value, bold=False, fill=None, align="left", wrap=False, font_color="000000", size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = fill
    c.border = BORDER
    return c


def generate_excel(assessment, pillars, responses, scores, level_names, comments=None):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    set_cell(ws, r, 1, "Safety Maturity Assessment Report", bold=True, fill=DARK_FILL, align="center", font_color="FFFFFF", size=13)
    ws.row_dimensions[r].height = 28

    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    set_cell(ws, r, 1, assessment["site_name"], bold=True, align="center", size=12)
    ws.row_dimensions[r].height = 22

    r += 1
    atype = assessment["type"].upper()
    scope = f" — {assessment['scope'].replace('_',' ').title()}" if assessment.get("scope") and assessment["type"] == "retail" else ""
    ws.merge_cells(f"A{r}:C{r}")
    set_cell(ws, r, 1, f"{atype}{scope}", align="center", font_color="6C757D")

    r += 2
    for label, value in [
        ("Assessment Date", assessment.get("assessment_date") or "—"),
        ("Master Assessor A", assessment.get("assessor_a") or "—"),
        ("Master Assessor B", assessment.get("assessor_b") or "—"),
    ]:
        set_cell(ws, r, 1, label, bold=True, fill=LIGHT_FILL)
        ws.merge_cells(f"B{r}:C{r}")
        set_cell(ws, r, 2, value)
        r += 1

    r += 1
    # Header
    for col, header in [(1, "Category"), (2, "Score"), (3, "Level")]:
        set_cell(ws, r, col, header, bold=True, fill=DARK_FILL, align="center", font_color="FFFFFF")
    ws.row_dimensions[r].height = 18
    r += 1

    ov = scores.get("overall")
    sa = scores.get("safety_awareness")
    si = scores.get("system_implementation")

    for label, val in [
        ("Overall Safety Maturity", ov),
        ("Safety Awareness (Leadership + TM Eng.)", sa),
        ("System Implementation (Org. + System)", si),
    ]:
        sc_str = f"{val:.2f}" if val else "—"
        lv_name = level_names.get(round(val) if val else 0, "—")
        fl = level_fill(val) if val else PatternFill("solid", fgColor="6C757D")
        set_cell(ws, r, 1, label, bold=True, fill=LIGHT_FILL)
        set_cell(ws, r, 2, sc_str, align="center", fill=fl, font_color="FFFFFF", bold=True)
        set_cell(ws, r, 3, lv_name, align="center", fill=fl, font_color="FFFFFF")
        r += 1

    r += 1
    set_cell(ws, r, 1, "Pillar Scores", bold=True)
    r += 1
    for pillar in pillars:
        ps = scores["pillars"].get(pillar["id"], {}).get("score")
        fl = level_fill(ps) if ps else PatternFill("solid", fgColor="6C757D")
        set_cell(ws, r, 1, f"  {pillar['name']}", fill=LIGHT_FILL)
        set_cell(ws, r, 2, str(ps) if ps is not None else "—", align="center", fill=fl, font_color="FFFFFF", bold=True)
        set_cell(ws, r, 3, level_names.get(ps, "—") if ps else "—", align="center", fill=fl, font_color="FFFFFF")
        r += 1

    # 1-point alerts
    alerts = []
    for pillar in pillars:
        for element in pillar["elements"]:
            es = scores["pillars"].get(pillar["id"], {}).get("elements", {}).get(element["id"])
            if es == 1:
                alerts.append(f"{pillar['name']} → {element['name']}")
    if alerts:
        r += 1
        ws.merge_cells(f"A{r}:C{r}")
        set_cell(ws, r, 1, "⚠ IMMEDIATE ACTION REQUIRED (Score = 1)", bold=True,
                 fill=PatternFill("solid", fgColor="DC3545"), align="center", font_color="FFFFFF")
        r += 1
        for a in alerts:
            ws.merge_cells(f"A{r}:C{r}")
            set_cell(ws, r, 1, f"• {a}", fill=PatternFill("solid", fgColor="F8D7DA"))
            r += 1

    # ── Sheet per pillar ──────────────────────────────────────────────────────
    for pillar in pillars:
        ws2 = wb.create_sheet(pillar["name"][:31])
        ws2.column_dimensions["A"].width = 5
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 50
        ws2.column_dimensions["D"].width = 16
        ws2.column_dimensions["E"].width = 35

        r2 = 1
        ws2.merge_cells(f"A{r2}:D{r2}")
        set_cell(ws2, r2, 1, pillar["name"], bold=True, fill=DARK_FILL, align="center", font_color="FFFFFF", size=12)
        ws2.row_dimensions[r2].height = 24

        ps = scores["pillars"].get(pillar["id"], {}).get("score")
        r2 += 1
        ws2.merge_cells(f"A{r2}:D{r2}")
        fl = level_fill(ps) if ps else PatternFill("solid", fgColor="6C757D")
        set_cell(ws2, r2, 1, f"Pillar Score: {ps} — {level_names.get(ps,'—')}" if ps else "Pillar Score: —",
                 bold=True, fill=fl, align="center", font_color="FFFFFF")

        r2 += 1
        for col, hdr in [(1,"No"),(2,"Element"),(3,"Question"),(4,"Answer"),(5,"Comment")]:
            set_cell(ws2, r2, col, hdr, bold=True, fill=LIGHT_FILL, align="center")

        for element in pillar["elements"]:
            es = scores["pillars"].get(pillar["id"], {}).get("elements", {}).get(element["id"])
            r2 += 1
            ws2.merge_cells(f"A{r2}:D{r2}")
            el_fl = level_fill(es) if es is not None else PatternFill("solid", fgColor="6C757D")
            set_cell(ws2, r2, 1,
                     f"{element['name']}  —  Score: {es} ({level_names.get(es,'N/A')})" if es is not None else f"{element['name']}  —  N/A",
                     bold=True, fill=el_fl, font_color="FFFFFF")
            ws2.row_dimensions[r2].height = 18

            for q in element["questions"]:
                ans = responses.get(q["id"], "")
                ans_label = ANSWER_LABELS.get(ans, "—")
                ans_hex = ANSWER_HEX.get(ans, "FFFFFF")
                ans_fill = PatternFill("solid", fgColor=ans_hex)
                cmt = (comments or {}).get(q["id"], "") or ""
                r2 += 1
                set_cell(ws2, r2, 1, q["no"], align="center")
                set_cell(ws2, r2, 2, element["name"], fill=LIGHT_FILL)
                set_cell(ws2, r2, 3, q.get("text") or "", wrap=True)
                set_cell(ws2, r2, 4, ans_label, align="center", fill=ans_fill, bold=(ans in ("yes","no")))
                set_cell(ws2, r2, 5, cmt, wrap=True)
                ws2.row_dimensions[r2].height = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
